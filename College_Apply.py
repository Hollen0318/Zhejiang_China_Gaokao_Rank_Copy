import openpyxl
import xlsxwriter
#import time
#import xml.dom.minidom
#import warnings

xlsm_file = openpyxl.load_workbook("Source.xlsx")
Ranking_Sheet = xlsm_file['Ranking']
workbook = xlsxwriter.Workbook('output.xlsx')
worksheet = workbook.add_worksheet()
School_Major = xlsm_file['School_Major']
Source_list = [0] * 9035
for row in range(2, School_Major.max_row+1, 1):
    if Source_list[int(School_Major["A{}".format(row)].value)] == 0:
        Source_list[int(School_Major["A{}".format(row)].value)] = [int(School_Major["B{}".format(row)].value)]
    else:
        Source_list[int(School_Major["A{}".format(row)].value)].append(int(School_Major["B{}".format(row)].value))

for row in range(2, Ranking_Sheet.max_row+1, 1):
    if Source_list[int(Ranking_Sheet["A{}".format(row)].value)] == 0:
        worksheet.write('B{}'.format(row-1), Ranking_Sheet["B{}".format(row)].value)
        worksheet.write('A{}'.format(row-1), Ranking_Sheet["A{}".format(row)].value)
        worksheet.write('C{}'.format(row-1), Ranking_Sheet["C{}".format(row)].value)
        worksheet.write('D{}'.format(row-1), Ranking_Sheet["D{}".format(row)].value)
        worksheet.write('E{}'.format(row-1), Ranking_Sheet["E{}".format(row)].value)
        worksheet.write('F{}'.format(row-1), Ranking_Sheet["F{}".format(row)].value)
        worksheet.write('G{}'.format(row-1), Ranking_Sheet["G{}".format(row)].value)
        worksheet.write('H{}'.format(row-1), "NO")
    else:
        if Source_list[int(Ranking_Sheet["A{}".format(row)].value)].count(int(Ranking_Sheet["C{}".format(row)].value)) > 0:
            worksheet.write('B{}'.format(row-1), Ranking_Sheet["B{}".format(row)].value)
            worksheet.write('A{}'.format(row-1), Ranking_Sheet["A{}".format(row)].value)
            worksheet.write('C{}'.format(row-1), Ranking_Sheet["C{}".format(row)].value)
            worksheet.write('D{}'.format(row-1), Ranking_Sheet["D{}".format(row)].value)
            worksheet.write('E{}'.format(row-1), Ranking_Sheet["E{}".format(row)].value)
            worksheet.write('F{}'.format(row-1), Ranking_Sheet["F{}".format(row)].value)
            worksheet.write('G{}'.format(row-1), Ranking_Sheet["G{}".format(row)].value)
            worksheet.write('H{}'.format(row-1), "YES")
        else:
            worksheet.write('B{}'.format(row-1), Ranking_Sheet["B{}".format(row)].value)
            worksheet.write('A{}'.format(row-1), Ranking_Sheet["A{}".format(row)].value)
            worksheet.write('C{}'.format(row-1), Ranking_Sheet["C{}".format(row)].value)
            worksheet.write('D{}'.format(row-1), Ranking_Sheet["D{}".format(row)].value)
            worksheet.write('E{}'.format(row-1), Ranking_Sheet["E{}".format(row)].value)
            worksheet.write('F{}'.format(row-1), Ranking_Sheet["F{}".format(row)].value)
            worksheet.write('G{}'.format(row-1), Ranking_Sheet["G{}".format(row)].value)
            worksheet.write('H{}'.format(row-1), "NO")

workbook.close()

#row_index = 0

#for row in range (1, max_row+1):
#   if (xlsm_product_sheet["A{}".format(row)].value == None):
#       break;
#   if (xlsm_product_sheet[])

#   打印专业