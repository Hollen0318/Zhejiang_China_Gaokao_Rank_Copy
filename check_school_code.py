#!/usr/bin/env python3

#检查学校编号是否有变化
import openpyxl
import xlsxwriter

xlsm_file = openpyxl.load_workbook("Source.xlsx")
School_Major = xlsm_file['School_Major']
Ranking_Sheet = xlsm_file['Ranking']
workbook = xlsxwriter.Workbook('School_Name_Check_Output.xlsx')
worksheet = workbook.add_worksheet()

Source_list = [0] * 9035
Source_list_2 = [0] * 9035
for row in range(2, School_Major.max_row+1, 1):
	if Source_list[int(School_Major["A{}".format(row)].value)] == 0:
		Source_list[int(School_Major["A{}".format(row)].value)] = School_Major["B{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")")
		
for row in range(2, Ranking_Sheet.max_row+1, 1):
	if Source_list_2[int(Ranking_Sheet["A{}".format(row)].value)] == 0:
		Source_list_2[int(Ranking_Sheet["A{}".format(row)].value)] = Ranking_Sheet["B{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")")
		
for i in range (1, 9035, 1):
	if	Source_list[i] != Source_list_2[i]:
		worksheet.write("A{}".format(i+1),i)
		worksheet.write("B{}".format(i+1),Source_list[i])
		worksheet.write("C{}".format(i+1),Source_list[i])
		worksheet.write("D{}".format(i+1),"Changed")
		print(str(Source_list[i]) + " " + str(Source_list_2[i]) + " Code " + str(i) + " is not consistent!")
	else:
		worksheet.write("A{}".format(i+1),i)
		worksheet.write("B{}".format(i+1),Source_list[i])
		worksheet.write("C{}".format(i+1),Source_list[i])
		worksheet.write("D{}".format(i+1),"Same")
		print(str(Source_list[i]) + " " + str(Source_list_2[i]) + " Code " + str(i) + " is consistent")
		
workbook.close()