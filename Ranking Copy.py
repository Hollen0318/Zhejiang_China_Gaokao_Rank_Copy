import openpyxl
import xlsxwriter
import time
#import xml.dom.minidom
#import warnings
starttime = time.time()
xlsm_file = openpyxl.load_workbook("Source.xlsx")
Ranking_Sheet = xlsm_file['Ranking']
workbook = xlsxwriter.Workbook('Ranking_Copy_Output.xlsx')
worksheet = workbook.add_worksheet()
School_Major = xlsm_file['School_Major']
Source_list = [0] * 9035

def findRnak(schooCdoe, majorName):
    for row in range(2, Ranking_Sheet.max_row+1, 1):
        if int(Ranking_Sheet["A{}".format(row)].value) == schooCdoe and Ranking_Sheet["D{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")") == majorName:
            if  Ranking_Sheet["G{}".format(row)].value == None:
                return 999999
            else:
                return int(Ranking_Sheet["G{}".format(row)].value)
for row in range(2, Ranking_Sheet.max_row+1, 1):
    if Source_list[int(Ranking_Sheet["A{}".format(row)].value)] == 0:
        Source_list[int(Ranking_Sheet["A{}".format(row)].value)] = [Ranking_Sheet["D{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")")]
    else:
        Source_list[int(Ranking_Sheet["A{}".format(row)].value)].append(Ranking_Sheet["D{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")"))

print("编号3216的学校2021年专业有 " + str(Source_list[3216]))
print("编号0002的学校2021年专业有 " + str(Source_list[2]))
print("编号3728的学校2021年专业有 " + str(Source_list[3728]))
print("编号4457的学校2021年专业有 " + str(Source_list[4457]))
find = 0
total = 0
for row in range(2, School_Major.max_row+1, 1):
    if Source_list[int(School_Major["A{}".format(row)].value)] == 0:
        worksheet.write('A{}'.format(row), School_Major["A{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")"))
        worksheet.write('B{}'.format(row), School_Major["B{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")"))
        worksheet.write('C{}'.format(row), School_Major["C{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")"))
        worksheet.write('D{}'.format(row), School_Major["D{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")"))
        worksheet.write('F{}'.format(row), School_Major["F{}".format(row)].value)
        worksheet.write('E{}'.format(row), School_Major["E{}".format(row)].value)
        worksheet.write('G{}'.format(row), School_Major["G{}".format(row)].value)
        worksheet.write('H{}'.format(row), School_Major["H{}".format(row)].value)
        worksheet.write('I{}'.format(row), School_Major["I{}".format(row)].value)
        worksheet.write('J{}'.format(row), School_Major["J{}".format(row)].value)
        worksheet.write('K{}'.format(row), School_Major["K{}".format(row)].value)
        worksheet.write('L{}'.format(row), School_Major["L{}".format(row)].value)
        worksheet.write('M{}'.format(row), "新校名")
        total += 1
    else:
        #       筛选出了2021 年排名有对齐的学校，也就是排名在范围内
        #       下一步是确认是否有专业在去年排名内，有则抄排名
        if Source_list[int(School_Major["A{}".format(row)].value)].count(School_Major["D{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")")) > 0:
            worksheet.write('A{}'.format(row), School_Major["A{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")"))
            worksheet.write('B{}'.format(row), School_Major["B{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")"))
            worksheet.write('C{}'.format(row), School_Major["C{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")"))
            worksheet.write('D{}'.format(row), School_Major["D{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")"))
            worksheet.write('F{}'.format(row), School_Major["F{}".format(row)].value)
            worksheet.write('E{}'.format(row), School_Major["E{}".format(row)].value)
            worksheet.write('G{}'.format(row), School_Major["G{}".format(row)].value)
            worksheet.write('H{}'.format(row), School_Major["H{}".format(row)].value)
            worksheet.write('I{}'.format(row), School_Major["I{}".format(row)].value)
            worksheet.write('J{}'.format(row), School_Major["J{}".format(row)].value)
            worksheet.write('K{}'.format(row), School_Major["K{}".format(row)].value)
            worksheet.write('L{}'.format(row), School_Major["L{}".format(row)].value)
            worksheet.write('M{}'.format(row), findRnak(int(School_Major["A{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")")), School_Major["D{}".format(row)].value))
            find += 1
            total += 1
        else:
            worksheet.write('A{}'.format(row), School_Major["A{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")"))
            worksheet.write('B{}'.format(row), School_Major["B{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")"))
            worksheet.write('C{}'.format(row), School_Major["C{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")"))
            worksheet.write('D{}'.format(row), School_Major["D{}".format(row)].value.replace(" ", "").replace("（","(").replace("）",")"))
            worksheet.write('F{}'.format(row), School_Major["F{}".format(row)].value)
            worksheet.write('E{}'.format(row), School_Major["E{}".format(row)].value)
            worksheet.write('G{}'.format(row), School_Major["G{}".format(row)].value)
            worksheet.write('H{}'.format(row), School_Major["H{}".format(row)].value)
            worksheet.write('I{}'.format(row), School_Major["I{}".format(row)].value)
            worksheet.write('J{}'.format(row), School_Major["J{}".format(row)].value)
            worksheet.write('K{}'.format(row), School_Major["K{}".format(row)].value)
            worksheet.write('L{}'.format(row), School_Major["L{}".format(row)].value)
            worksheet.write('M{}'.format(row), "新专业名")
            total += 1
            
worksheet.write("A1","院校代码")
worksheet.write("B1","院校名称")
worksheet.write("C1","专业代码")
worksheet.write("D1","专业名称")
worksheet.write("E1","学制")
worksheet.write("F1","省")
worksheet.write("G1","城市")
worksheet.write("H1","本专科")
worksheet.write("I1","计划数")
worksheet.write("J1","选考科目要求")
worksheet.write("K1","收费标准")
worksheet.write("L1","备注")
worksheet.write("M1","排名")
worksheet.write("N1","生物地理技术 2022 年录取专业的 2021 年分数线排名抄写成功")
a = "2021年专业找到排名 " + str(find)  + "条， 总共共 " + str(total) + " 条。"
endtime = time.time()
totalSeconds = format(endtime - starttime, ".3f")
b = "\n共耗时{}秒".format(totalSeconds)
worksheet.write("O1",a)
worksheet.write("P1",b)
workbook.close()

print("生物地理技术 2022 年录取专业的 2021 年分数线排名抄写成功")
print("共计2021年专业找到排名 " + str(find)  + "条， 总共共 " + str(total))
print("共耗时{}秒".format(totalSeconds))